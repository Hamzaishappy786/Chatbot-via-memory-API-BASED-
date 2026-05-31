from openpyxl import load_workbook
from app.ingestion.parsers.base import BaseParser, ParsedContent


class XLSXParser(BaseParser):
    @staticmethod
    def supported_extensions() -> list[str]:
        return [".xlsx", ".xlsm"]

    def parse(self, file_path: str) -> ParsedContent:
        wb = load_workbook(file_path, data_only=True)
        content = ParsedContent()

        for sheet_num, sheet_name in enumerate(wb.sheetnames, 1):
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))

            if rows:
                header = rows[0]
                col_count = len(list(ws.iter_cols(max_row=1)))
                separator = " | ".join(["---"] * col_count)
                table_md = f"Sheet: {sheet_name}\n" + "\n".join([header, separator] + rows[1:])
                content.text_blocks.append({
                    "text": table_md,
                    "page": sheet_num,
                })

        wb.close()
        return content
